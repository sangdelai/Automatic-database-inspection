#**
#Mysql auto check V1.1
#help bro more and more lazy~
#@author:huminghui
#@data:2024-01-29
#**

try:
    from openpyxl import load_workbook
    import pymysql
    import paramiko
    import re
except:
    print('请安装openpyxl、pymysql、paramiko库后运行此程序，如果不会请联系胡明慧付费咨询')


data = input("请输入巡检的月份（例如1月输入数字1）：")



def mysql_check(ip,port,username,password,sqluser,sqlpasswd):
    data_every = []
    data_do_not_have = []
    # for eachday in range(1,32):
    #     datas = '2023-' + str(data) + '-' + str(eachday)
    #     data_every.append(datas)
    # print(data_every)
    datas = '2023-' + str(data) + '-'
    # host = '192.168.14.52'
    ssh = paramiko.SSHClient()
    # 允许将信任的主机自动加入到host_allow 列表，此方法必须放在connect方法的前面
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    # 调用connect方法连接服务器
    ssh.connect(hostname=ip, port=port, username=username, password=password)
    # 执行命令
    stdin, stdout, stderr = ssh.exec_command('df -h /')
    # 结果放到stdout中，如果有错误将放到stderr中
    output = stdout.read().decode()
    result_df = re.findall(r'  (.*?)G',output)
    result_df = float(result_df[0])
    # print(result_df)
    # 关闭连接
    # ssh.close()

    try:
        print("正在巡检mysql数据库"+ip)
        conn = pymysql.connect(host=ip,user=sqluser,passwd=sqlpasswd,database='information_schema')
        cursor = conn.cursor()
        cursor.execute('''SELECT table_schema "数据库名", ROUND(SUM(data_length + index_length) / 1024 / 1024 / 1024,2) "占用磁盘空间(GB)" FROM information_schema.TABLES;''')
        result = cursor.fetchall()
        with open('result.txt', 'a') as f:
            f.write('检查项1:检查空间使用率是否大于90%\n')
        print('检查项1:检查空间使用率是否大于90%')
        for row in result:
            with open('result.txt', 'a') as f:
                f.write("磁盘占用率："+str(round(float(row[1])/result_df,2))+'\n')
            print("磁盘占用率："+str(round(float(row[1])/result_df,2)))
    except:
        with open('result.txt', 'a') as f:
            f.write('磁盘占用率功能存在问题，请联系开发者。**\n')
        print('磁盘占用率功能存在问题，请联系开发者。**')


    cursor.execute('''show open tables where in_use>0;''')
    result_lock = cursor.fetchall()
    # print(result_lock[0])
    # print(list(result_lock))
    with open('result.txt', 'a') as f:
        f.write('检查项2:是否存在死锁问题\n')
    print('检查项2:是否存在死锁问题')
    if list(result_lock) == []:
        with open('result.txt', 'a') as f:
            f.write("无死锁\n")
        print("无死锁")
    else:
        with open('result.txt', 'a') as f:
            f.write('死锁存在问题'+str(result_lock)+'**\n')
        print('死锁存在问题'+str(result_lock)+'**')


    cursor.execute('''show variables like 'max_connections';''')
    result_max = cursor.fetchall()
    cursor.execute('''show status like 'Threads_connected';''')
    result_threads = cursor.fetchall()
    with open('result.txt', 'a') as f:
        f.write('检查项3:最大连接数是否大于当前连接数\n')
    print('检查项3:最大连接数是否大于当前连接数')
    if int(result_max[0][1])>int(result_threads[0][1]):
        with open('result.txt', 'a') as f:
            f.write('resutl_max的值为：'+result_max[0][1]+'  ' +'result_threads的值为'+result_threads[0][1]+'  ' +  'resutl_max>result_threads' +"连接数没有问题\n")
        print('resutl_max的值为：'+result_max[0][1]+'  ' +'result_threads的值为'+result_threads[0][1]+'  ' +  'resutl_max>result_threads' +"连接数没有问题")
    else:
        with open('result.txt', 'a') as f:
            f.write('resutl_max的值为：'+result_max[0][1]+'  ' +'result_threads的值为'+result_threads[0][1]+'  ' + 'resutl_max<=result_threads'+'连接数存在问题，请检查**\n')
        print('resutl_max的值为：'+result_max[0][1]+'  ' +'result_threads的值为'+result_threads[0][1]+'  ' + 'resutl_max<=result_threads'+'连接数存在问题，请检查**')
    # print(result_max[0][1])
    # print(result_threads[0][1])


    cursor.execute('''show global status like 'Innodb_buffer_pool_read_requests';''')
    result_buffer_pool_read_requests = cursor.fetchall()
    # print(result_buffer_pool_read_requests[0][1])
    cursor.execute('''show global status like 'Innodb_buffer_pool_reads';''')
    result_buffer_pool_reads = cursor.fetchall()
    with open('result.txt', 'a') as f:
        f.write('检查项4:检查InnoDB buffer命中概率是否大于90% \n')
    print('检查项4:检查InnoDB buffer命中概率是否大于90% ')
    # print(result_buffer_pool_reads[0][1])
    result_buffer = round(1 - (float(result_buffer_pool_reads[0][1])/float(result_buffer_pool_read_requests[0][1])),5)
    if result_buffer > 0.9:
        with open('result.txt', 'a') as f:
            f.write('Innodb_buffer_pool_read_requests的值为：'+result_buffer_pool_read_requests[0][1]+'  ' +'result_buffer_pool_reads的值为：'+result_buffer_pool_reads[0][1] +'\n'+'1-Innodb_buffer_pool_read_requests/Innodb_buffer_pool_reads的值为'+str(result_buffer)+ "  InnoDB buffer命中率没有问题\n")
        print('Innodb_buffer_pool_read_requests的值为：'+result_buffer_pool_read_requests[0][1]+'  ' +'result_buffer_pool_reads的值为：'+result_buffer_pool_reads[0][1] +'\n'+'1-Innodb_buffer_pool_read_requests/Innodb_buffer_pool_reads的值为'+str(result_buffer)+ "  InnoDB buffer命中率没有问题")
    else:
        with open('result.txt', 'a') as f:
            f.write('Innodb_buffer_pool_read_requests的值为：'+result_buffer_pool_read_requests[0][1]+'  ' +'result_buffer_pool_reads的值为：'+result_buffer_pool_reads[0][1] +'\n'+"InnoDB buffer命中率有问题," +'1-Innodb_buffer_pool_read_requests/Innodb_buffer_pool_reads的值为'+ str(result_buffer)+'**\n')
        print('Innodb_buffer_pool_read_requests的值为：'+result_buffer_pool_read_requests[0][1]+'  ' +'result_buffer_pool_reads的值为：'+result_buffer_pool_reads[0][1] +'\n'+"InnoDB buffer命中率有问题," +'1-Innodb_buffer_pool_read_requests/Innodb_buffer_pool_reads的值为'+ str(result_buffer)+'**')
    # print("InnoDB buffer命中率为："+ str(result_buffer))



    cursor.execute('''show global status like 'key_read_requests';''')
    result_key_read_requests = cursor.fetchall()
    # print(float(result_key_read_requests[0][1]))
    cursor.execute('''show global status like 'key_reads';''')
    result_key_read = cursor.fetchall()
    with open('result.txt', 'a') as f:
        f.write('检查项5:检查key buffer命中概率是否大于85%\n')
    print('检查项5:检查key buffer命中概率是否大于85%')
    # print(float(result_key_read[0][1]))

    try:
        result_key = round(1 - (float(result_key_read[0][1])/float(result_key_read_requests[0][1])),5)
        if result_key > 0.85:
            with open('result.txt', 'a') as f:
                f.write('key_read_requests的值为' + str(float(result_key_read_requests[0][1])) + '  ' + 'key_reads的值为：' + str(float(result_key_read[0][1])) + '\n' + '1 - result_key_read/result_key_read_requests的值为：' + str(result_key)+'  '+"key buffer命中率没有问题\n")
            print('key_read_requests的值为' + str(float(result_key_read_requests[0][1])) + '  ' + 'key_reads的值为：' + str(float(result_key_read[0][1])) + '\n' + '1 - result_key_read/result_key_read_requests的值为：' + str(result_key)+'  '+"key buffer命中率没有问题")
        else:
            with open('result.txt', 'a') as f:
                f.write('key_read_requests的值为' + str(float(result_key_read_requests[0][1])) + '  ' + 'key_reads的值为：' + str(float(result_key_read[0][1])) + '\n' + '1 - result_key_read/result_key_read_requests的值为：' + str(result_key)+'  '+"key buffer命中率有问题\n")
            print('key_read_requests的值为' + str(float(result_key_read_requests[0][1])) + '  ' + 'key_reads的值为：' + str(float(result_key_read[0][1])) + '\n' + '1 - result_key_read/result_key_read_requests的值为：' + str(result_key)+'  '+"key buffer命中率有问题")
    except:
        with open('result.txt', 'a') as f:
            f.write('key buffer存在问题为0'+'**\n')
        print('key buffer存在问题为0'+'**')

    cursor.execute('''show global status like 'Innodb_log_waits';''')
    result_Innodb_log_waits = cursor.fetchall()
    with open('result.txt', 'a') as f:
        f.write('检查项6:检查Innodb_log_waits量\n')
    print('检查项6:检查Innodb_log_waits量')
    if result_Innodb_log_waits[0][1] == '0':
        with open('result.txt', 'a') as f:
            f.write('Innodb_log_waits的值为：'+result_Innodb_log_waits[0][1]+'  '+'Innodb_log_waits量没有问题\n')
        print('Innodb_log_waits的值为：'+result_Innodb_log_waits[0][1]+'  '+'Innodb_log_waits量没有问题')
    else:
        with open('result.txt', 'a') as f:
            f.write('Innodb_log_waits量有问题，问题为：'+result_Innodb_log_waits[0][1]+'**\n')
        print('Innodb_log_waits量有问题，问题为：'+result_Innodb_log_waits[0][1]+'**')


    cursor.execute('''show global status like 'Binlog_cache_disk_use';''')
    result_Binlog_Cache = cursor.fetchall()
    print('检查项7:检查Binlog Cache使用情况是否小于3')
    with open('result.txt', 'a') as f:
        f.write('检查项7:检查Binlog Cache使用情况是否小于3\n')
    # print(result_Binlog_Cache[0][1])
    if int(result_Binlog_Cache[0][1]) < 3:
        with open('result.txt', 'a') as f:
            f.write('Binlog_cache_disk_use的值为：'+result_Binlog_Cache[0][1]+'  '+"result_Binlog Cache没问题\n")
        print('Binlog_cache_disk_use的值为：'+result_Binlog_Cache[0][1]+'  '+"result_Binlog Cache没问题")
    else:
        with open('result.txt', 'a') as f:
            f.write("Binlog Cache有问题，Binlog Cache值为"+ str(result_Binlog_Cache[0][1])+'**\n')
        print("Binlog Cache有问题，Binlog Cache值为"+ str(result_Binlog_Cache[0][1])+'**')


    cursor.execute('''show global status like 'Table_open_cache_hits';''')
    result_Table_open_cache_hits = cursor.fetchall()
    # print(result_Table_open_cache_hits)
    cursor.execute('''show global status like 'Table_open_cache_misses';''')
    result_Table_open_cache_misses = cursor.fetchall()
    with open('result.txt', 'a') as f:
        f.write('检查项8:检查table cache状态量是否大于85%\n')
    print('检查项8:检查table cache状态量是否大于85%')
    # print(result_Table_open_cache_misses)
    if float(int(result_Table_open_cache_hits[0][1])/(int(result_Table_open_cache_hits[0][1])+int(result_Table_open_cache_misses[0][1]))) > 0.85:
        with open('result.txt', 'a') as f:
            f.write('Table_open_cache_hits的值为：'+result_Table_open_cache_hits[0][1]+'  '+'result_Table_open_cache_hits的值为：'+result_Table_open_cache_hits[0][1] +'\n'+'Table_open_cache_hits/(Table_open_cache_misses+Table_open_cache_hits)的值为：'+str(float(int(result_Table_open_cache_hits[0][1])/(int(result_Table_open_cache_hits[0][1])+int(result_Table_open_cache_misses[0][1]))))+'  '+'table cache状态量没问题\n')
        print('Table_open_cache_hits的值为：'+result_Table_open_cache_hits[0][1]+'  '+'result_Table_open_cache_hits的值为：'+result_Table_open_cache_hits[0][1] +'\n'+'Table_open_cache_hits/(Table_open_cache_misses+Table_open_cache_hits)的值为：'+str(float(int(result_Table_open_cache_hits[0][1])/(int(result_Table_open_cache_hits[0][1])+int(result_Table_open_cache_misses[0][1]))))+'  '+'table cache状态量没问题')
    else:
        with open('result.txt', 'a') as f:
            f.write('Table_open_cache_hits的值为：'+result_Table_open_cache_hits[0][1]+'  '+'result_Table_open_cache_hits的值为：'+result_Table_open_cache_hits[0][1] +'\n'+'Table_open_cache_hits/(Table_open_cache_misses+Table_open_cache_hits)的值为：'+ str(round(float(result_Table_open_cache_hits[0][1])/(float(result_Table_open_cache_hits[0][1])+float(result_Table_open_cache_misses[0][1])),2))+'table cache状态量有问题**\n')
        print('Table_open_cache_hits的值为：'+result_Table_open_cache_hits[0][1]+'  '+'result_Table_open_cache_hits的值为：'+result_Table_open_cache_hits[0][1] +'\n'+'Table_open_cache_hits/(Table_open_cache_misses+Table_open_cache_hits)的值为：'+ str(round(float(result_Table_open_cache_hits[0][1])/(float(result_Table_open_cache_hits[0][1])+float(result_Table_open_cache_misses[0][1])),2))+'table cache状态量有问题**')


    cursor.execute('''SHOW STATUS LIKE 'Threads_created';  ''')
    result_Threads_created = cursor.fetchall()
    # print(result_Threads_created)
    cursor.execute('''show global status like 'Connections';''')
    result_Connections = cursor.fetchall()
    with open('result.txt', 'a') as f:
        f.write('检查项9:检查Thread cache命中率是否大于90%\n')
    print('检查项9:检查Thread cache命中率是否大于90%')
    # print(result_Connections)
    result_Thread_cache = round(1 - (float(result_Threads_created[0][1])/float(result_Connections[0][1])),5)
    if result_Thread_cache > 0.9:
        with open('result.txt', 'a') as f:
            f.write('Threads_created的值为：'+result_Threads_created[0][1]+'  '+'Connections的值为：'+result_Connections[0][1]+'\n'+'1-(Threads_created / Connections)的值为：'+str(result_Thread_cache)+'  '+"Thread cache命中率没有问题\n")

        print('Threads_created的值为：'+result_Threads_created[0][1]+'  '+'Connections的值为：'+result_Connections[0][1]+'\n'+'1-(Threads_created / Connections)的值为：'+str(result_Thread_cache)+'  '+"Thread cache命中率没有问题")
    else:
        with open('result.txt', 'a') as f:
            f.write('Threads_created的值为：'+result_Threads_created[0][1]+'  '+'Connections的值为：'+result_Connections[0][1]+'\n'+'1-(Threads_created / Connections)的值为：'+str(result_Thread_cache)+'  '+"Thread cache命中率有问题**\n")
        print('Threads_created的值为：'+result_Threads_created[0][1]+'  '+'Connections的值为：'+result_Connections[0][1]+'\n'+'1-(Threads_created / Connections)的值为：'+str(result_Thread_cache)+'  '+"Thread cache命中率有问题**")


    cursor.execute('''show variables like 'Slow_queries';''')
    result_queries = cursor.fetchall()
    # print(result_queries)
    if list(result_queries) == []:
        with open('result.txt', 'a') as f:
            f.write('慢查询记录没问题，无新增\n')
        print('慢查询记录没问题，无新增')
    else:
        with open('result.txt', 'a') as f:
            f.write('慢查询记录存在新增，新增内容为：'+ result_queries+'**\n')
        print('慢查询记录存在新增，新增内容为：'+ result_queries+'**')



    cursor.execute('''SHOW VARIABLES LIKE 'log_error';''')
    result_log_error = cursor.fetchall()
    result_log_error = str(result_log_error[0][1])
    stdin, stdout, stderr = ssh.exec_command('cat ' + result_log_error)
    # 结果放到stdout中，如果有错误将放到stderr中
    output2 = stdout.read().decode()
    # print(output2)
    # print(re.findall(r'\[ERROR](.*?)\n',output2))

    pattern = r'^.*\[ERROR].*$'
    # 使用re.MULTILINE标志以允许^和$匹配每一行的开始和结束
    match = re.findall(pattern, output2, re.MULTILINE)
    # print(match)
    # 如果匹配成功，则打印整行
    try:
        if match == []:
            print("没有ERROR")
        else:
            for i in match:
                # print(i)
                if datas in i:
                    print(i)
                    with open('result.txt', 'a') as f:
                        f.write(i+'\n')
                else:
                    data_do_not_have.append('当月无REEOR错误')
                    with open('result.txt', 'a') as f:
                        f.write('当月无REEOR错误\n')


    except:
        print('没有ERROR报错！')
        with open('result.txt', 'a') as f:
            f.write('没有ERROR报错！\n')
    try:
        print(data_do_not_have[0])
    except:
        pass


    ssh.close()
    cursor.close()
    conn.close()




def check_and_get_dev_list(filename, sheet_name):
    excel_information = []
    sheet_header = []
    wb = load_workbook(filename)
    sh = wb[sheet_name]
    # 获取最大行数
    row = sh.max_row
    # 获取最大列数
    column = sh.max_column
    data = []
    # 获取表头写入列表中方便调用
    for data_1 in range(1, column + 1):
        get_sheet_header = sh.cell(row=1, column=data_1).value
        sheet_header.append(get_sheet_header)
    # 第一行为表头， 此处 row +1 是pyton循环时不读取最后一个数
    for row_1 in range(2, row + 1):
        # 存储一行信息
        sheet_data_1 = dict()
        # 逐行读取表中的数据
        for b in range(1, column + 1):
            cell = sh.cell(row=row_1, column=b).value
            # 将数据已字典形式写入 sheet_data_1 中
            # if cell != None:
            sheet_data_1[sheet_header[b - 1]] = cell
        excel_information.append(sheet_data_1)
    for i in excel_information:
        if i['ip'] != None:
            data.append(i)
    return data


res = check_and_get_dev_list('./resource.xlsx', 'Sheet1')
# print(res)
for eachone in res:
    # print(eachone)
    # print(eachone['ip'])
    ip = eachone['ip']
    port = eachone['port']
    username = eachone['username']
    password = eachone['password']
    sqlsuer = eachone['sqluser']
    sqlpasswd = eachone['sqlpasswd']
    mysql_check(ip,port,username,password,sqlsuer,sqlpasswd)
    print('--------------------------------------------------')
    with open('result.txt', 'a') as f:
        f.write('----------------------------------------------------------\n')
# print(res)

